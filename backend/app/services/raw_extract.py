"""从多种 raw 格式提取纯文本（供侧车 .extract.md 与后续 ingest）。"""

from __future__ import annotations

import csv
import json
import re
from io import BytesIO, StringIO
from pathlib import Path

# 允许上传的扩展名（小写）
ALLOWED_RAW_SUFFIXES = frozenset(
    {
        ".md",
        ".markdown",
        ".txt",
        ".csv",
        ".json",
        ".html",
        ".htm",
        ".pdf",
        ".docx",
        ".xlsx",
        ".xlsm",
        ".xls",
    }
)

# 保存二进制原件，并尝试生成 .extract.md
EXTRACTABLE_SUFFIXES = frozenset({".pdf", ".docx", ".xlsx", ".xlsm", ".xls"})

_TEXT_SUFFIXES = ALLOWED_RAW_SUFFIXES - EXTRACTABLE_SUFFIXES


def is_allowed_suffix(suffix: str) -> bool:
    return suffix.lower() in ALLOWED_RAW_SUFFIXES


def extract_sidecar_rel(raw_rel: str) -> str:
    p = Path(raw_rel)
    return (p.parent / f"{p.stem}.extract.md").as_posix()


def extract_text(filename: str, data: bytes) -> tuple[str | None, str]:
    """
    从文件字节提取文本。
    返回 (text, note)；text 为 None 表示无法提取（仍保留原件）。
    """
    suffix = Path(filename).suffix.lower()

    if suffix in _TEXT_SUFFIXES:
        try:
            return data.decode("utf-8"), "utf-8 文本"
        except UnicodeDecodeError:
            return data.decode("utf-8", errors="replace"), "utf-8 文本（含替换字符）"

    if suffix == ".pdf":
        return _extract_pdf(data)
    if suffix == ".docx":
        return _extract_docx(data)
    if suffix in (".xlsx", ".xlsm"):
        return _extract_xlsx(data)
    if suffix == ".xls":
        return _extract_xls(data)

    return None, "不支持的格式"


def _extract_pdf(data: bytes) -> tuple[str | None, str]:
    try:
        import fitz  # pymupdf
    except ImportError:
        return None, "未安装 pymupdf，请 pip install pymupdf"

    try:
        doc = fitz.open(stream=data, filetype="pdf")
        parts: list[str] = []
        for i, page in enumerate(doc):
            parts.append(f"## 第 {i + 1} 页\n\n{page.get_text().strip()}")
        doc.close()
        text = "\n\n".join(p for p in parts if p.strip())
        return text or None, f"PDF {len(parts)} 页"
    except Exception as exc:
        return None, f"PDF 解析失败: {exc}"


def _extract_docx(data: bytes) -> tuple[str | None, str]:
    try:
        from docx import Document
    except ImportError:
        return None, "未安装 python-docx，请 pip install python-docx"

    try:
        doc = Document(BytesIO(data))
        paras = [p.text.strip() for p in doc.paragraphs if p.text.strip()]
        text = "\n\n".join(paras)
        return text or None, f"Word {len(paras)} 段"
    except Exception as exc:
        return None, f"Word 解析失败: {exc}"


def _extract_xlsx(data: bytes) -> tuple[str | None, str]:
    try:
        import openpyxl
    except ImportError:
        return None, "未安装 openpyxl，请 pip install openpyxl"

    try:
        wb = openpyxl.load_workbook(BytesIO(data), read_only=True, data_only=True)
        parts: list[str] = []
        for sheet in wb.worksheets:
            rows: list[str] = []
            for row in sheet.iter_rows(values_only=True):
                cells = [str(c).strip() if c is not None else "" for c in row]
                if any(cells):
                    rows.append(" | ".join(cells))
            if rows:
                parts.append(f"## {sheet.title}\n\n" + "\n".join(rows))
        wb.close()
        text = "\n\n".join(parts)
        return text or None, f"Excel {len(wb.sheetnames)} 表"
    except Exception as exc:
        return None, f"Excel 解析失败: {exc}"


def _extract_xls(data: bytes) -> tuple[str | None, str]:
    try:
        import xlrd
    except ImportError:
        return None, "未安装 xlrd，请 pip install xlrd"

    try:
        book = xlrd.open_workbook(file_contents=data)
        parts: list[str] = []
        for sheet in book.sheets():
            rows: list[str] = []
            for rx in range(sheet.nrows):
                cells = [str(sheet.cell_value(rx, cx)).strip() for cx in range(sheet.ncols)]
                if any(cells):
                    rows.append(" | ".join(cells))
            if rows:
                parts.append(f"## {sheet.name}\n\n" + "\n".join(rows))
        text = "\n\n".join(parts)
        return text or None, f"Excel {book.nsheets} 表"
    except Exception as exc:
        return None, f"Excel(.xls) 解析失败: {exc}"


def format_csv_as_markdown(text: str) -> str:
    """将 CSV 文本格式化为简易 Markdown 表。"""
    reader = csv.reader(StringIO(text))
    rows = list(reader)
    if not rows:
        return text
    lines = []
    for i, row in enumerate(rows[:500]):
        line = "| " + " | ".join(cell.replace("|", "\\|") for cell in row) + " |"
        lines.append(line)
        if i == 0:
            lines.append("| " + " | ".join("---" for _ in row) + " |")
    if len(rows) > 500:
        lines.append(f"\n… 共 {len(rows)} 行，已截断")
    return "\n".join(lines)


def normalize_text_content(filename: str, text: str) -> str:
    suffix = Path(filename).suffix.lower()
    if suffix == ".csv":
        return format_csv_as_markdown(text)
    if suffix == ".json":
        try:
            obj = json.loads(text)
            return "```json\n" + json.dumps(obj, ensure_ascii=False, indent=2) + "\n```"
        except json.JSONDecodeError:
            return text
    if suffix in (".html", ".htm"):
        t = re.sub(r"<script[\s\S]*?</script>", "", text, flags=re.I)
        t = re.sub(r"<style[\s\S]*?</style>", "", t, flags=re.I)
        t = re.sub(r"<[^>]+>", " ", t)
        t = re.sub(r"\s+", " ", t).strip()
        return t
    return text
